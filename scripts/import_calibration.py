from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
REPORT = ROOT / "docs" / "CALIBRATION_REPORT.md"


def inspect_workbook(path: Path) -> str:
    book = load_workbook(path, read_only=True, data_only=True)
    lines = [f"## {path.name}", "", f"Sheets: {', '.join(book.sheetnames)}", ""]
    for sheet in book.sheetnames[:5]:
        ws = book[sheet]
        sample = []
        for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
            sample.append([str(cell)[:50] if cell is not None else "" for cell in row[:8]])
        lines.append(f"### {sheet}")
        lines.append("")
        for row in sample:
            lines.append(f"- {row}")
        lines.append("")
    return "\n".join(lines)


def main() -> None:
    expected = [
        ROOT / "Microskin_India_Territory_Analysis_v5.xlsx",
        ROOT / "Microskin_USA_Territory_Analysis_v3.xlsx",
    ]
    lines = [
        "# Calibration Report",
        "",
        "This report maps detected spreadsheet assets into the new engine's configuration and notes missing calibration sources.",
        "",
    ]
    for workbook in expected:
        if workbook.exists():
            lines.append(inspect_workbook(workbook))
        else:
            lines.extend(
                [
                    f"## {workbook.name}",
                    "",
                    "Status: not found in workspace during sandbox build.",
                    "Mapping approach: importer path reserved; country defaults and commercial assumptions seeded from transparent demo configuration until workbook is supplied.",
                    "",
                ]
            )
    REPORT.write_text("\n".join(lines), encoding="utf-8")


if __name__ == "__main__":
    main()

